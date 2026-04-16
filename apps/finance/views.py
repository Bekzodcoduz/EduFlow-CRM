import calendar
from datetime import date, timedelta
from io import BytesIO

from django.http import Http404, HttpResponseBadRequest
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.db.models import Count, Q, Sum
from .models import Payment
from apps.groups.models import Group, Course
from apps.students.models import Attendance, Student


def _ensure_admin(request):
    if request.user.is_admin:
        return None
    messages.error(request, "Moliyaviy bo'lim faqat administrator uchun.")
    return redirect('dashboard')


@login_required
def finance_list(request):
    admin_redirect = _ensure_admin(request)
    if admin_redirect:
        return admin_redirect

    tab = request.GET.get('tab', 'payments')

    payments = Payment.objects.select_related(
        'student__group', 'student__group__course', 'received_by', 'course', 'group'
    )
    month    = request.GET.get('month', '')
    if month:
        y, m = month.split('-')
        payments = payments.filter(created_at__year=y, created_at__month=m)

    total_income   = int(payments.filter(payment_type='income').aggregate(s=Sum('amount'))['s'] or 0)
    total_refund   = int(payments.filter(payment_type='refund').aggregate(s=Sum('amount'))['s'] or 0)
    total_discount = int(payments.filter(payment_type='discount').aggregate(s=Sum('amount'))['s'] or 0)
    today = date.today()
    month_start = today.replace(day=1)
    month_end = today.replace(day=calendar.monthrange(today.year, today.month)[1])
    all_active_students = Student.objects.filter(is_active=True).select_related('group')
    paid_student_ids = set(
        Payment.objects.filter(
            payment_type=Payment.INCOME,
            student__is_active=True,
            created_at__date__gte=month_start,
            created_at__date__lte=month_end,
        ).values_list('student_id', flat=True).distinct()
    )
    debtors = all_active_students.exclude(pk__in=paid_student_ids)
    debtors_total = debtors.count()

    from django.db.models.functions import TruncMonth
    monthly = (Payment.objects
               .filter(payment_type='income',
                       created_at__gte=date.today().replace(day=1) - timedelta(days=210))
               .annotate(month=TruncMonth('created_at'))
               .values('month')
               .annotate(total=Sum('amount'))
               .order_by('month'))
    monthly_data = [{'label': d['month'].strftime('%b'), 'total': int(d['total'])} for d in monthly]

    ctx = {
        'tab':           tab,
        'payments':      payments,
        'total_income':  total_income,
        'total_refund':  total_refund,
        'total_discount':total_discount,
        'debtors':       debtors,
        'debtors_total': debtors_total,
        'status_month_label': f'{today.month:02d}.{today.year}',
        'students':      Student.objects.filter(is_active=True).select_related('group', 'group__course').order_by(
            'last_name', 'first_name'
        ),
        'payment_courses': Course.objects.all().order_by('name'),
        'payment_groups': Group.objects.filter(is_active=True).select_related('course').order_by('name'),
        'export_groups': Group.objects.filter(is_active=True).order_by('name'),
        'monthly_data':  monthly_data,
        'month':         month,
    }
    return render(request, 'finance/list.html', ctx)


@login_required
@require_POST
def payment_create(request):
    admin_redirect = _ensure_admin(request)
    if admin_redirect:
        return admin_redirect

    sid    = request.POST.get('student')
    amount = request.POST.get('amount', '0')
    ptype  = request.POST.get('payment_type', 'income')
    if not sid or not amount:
        messages.error(request, "Talaba va summa majburiy!")
        nu = (request.POST.get('next') or '').strip()
        if nu.startswith('/'):
            return redirect(nu)
        return redirect('finance')

    course_raw = (request.POST.get('course') or '').strip()
    group_raw = (request.POST.get('group') or '').strip()
    course_id = None
    group_id = None
    if course_raw.isdigit():
        cid = int(course_raw)
        if Course.objects.filter(pk=cid).exists():
            course_id = cid

    if group_raw.isdigit():
        gid = int(group_raw)
        grp = Group.objects.filter(pk=gid).select_related('course').first()
        if grp:
            group_id = gid
            if course_id and grp.course_id and grp.course_id != course_id:
                messages.error(request, "Tanlangan guruh fan bilan mos emas.")
                nu = (request.POST.get('next') or '').strip()
                if nu.startswith('/'):
                    return redirect(nu)
                return redirect('finance')
            if not course_id and grp.course_id:
                course_id = grp.course_id

    method = (request.POST.get('payment_method') or Payment.METHOD_CASH).strip()
    valid_methods = {Payment.METHOD_CASH, Payment.METHOD_TRANSFER, Payment.METHOD_TERMINAL}
    if method not in valid_methods:
        method = Payment.METHOD_CASH

    Payment.objects.create(
        student_id   = sid,
        course_id    = course_id,
        group_id     = group_id,
        payment_method = method,
        payment_type = ptype,
        amount       = int(amount),
        note         = request.POST.get('note', ''),
        received_by  = request.user,
    )
    messages.success(request, "✓ To'lov saqlandi!")
    next_url = request.POST.get('next', '')
    if next_url and next_url.startswith('/'):
        return redirect(next_url)
    return redirect('finance')


@login_required
@require_POST
def payment_delete(request, pk):
    admin_redirect = _ensure_admin(request)
    if admin_redirect:
        return admin_redirect

    get_object_or_404(Payment, pk=pk).delete()
    messages.success(request, "O'chirildi!")
    return redirect('finance')


def _payment_export_queryset(student_pk: str | None, group_pk: str | None):
    """To'lovlar: ixtiyoriy guruh va/yoki talaba filtri."""
    qs = Payment.objects.select_related('student__group', 'received_by', 'course', 'group').order_by('-created_at')
    one = None
    grp = None
    if group_pk:
        grp = get_object_or_404(Group, pk=group_pk, is_active=True)
        qs = qs.filter(student__group=grp)
    if student_pk:
        one = get_object_or_404(Student, pk=student_pk, is_active=True)
        if grp is not None and one.group_id != grp.pk:
            raise Http404('Talaba tanlangan guruhda emas.')
        qs = qs.filter(student=one)
    return qs, one, grp


def _export_filename_suffix(one_student: Student | None, export_group: Group | None) -> str:
    parts = []
    if export_group:
        parts.append(f'guruh{export_group.pk}')
    if one_student:
        parts.append(f'talaba{one_student.pk}')
    return '_' + '_'.join(parts) if parts else ''


def _annotated_students_export_qs():
    return (
        Student.objects.filter(is_active=True)
        .select_related('group')
        .annotate(
            att_n=Count('attendances'),
            att_ok=Count('attendances', filter=Q(attendances__is_present=True)),
            att_ex=Count(
                'attendances',
                filter=Q(attendances__is_present=False, attendances__excused_absence=True),
            ),
            att_nopay=Count(
                'attendances',
                filter=Q(attendances__is_present=False, attendances__excused_absence=False),
            ),
        )
        .order_by('last_name', 'first_name')
    )


def _attendance_summary_rows_for_students_qs(students_qs):
    out = []
    for s in students_qs:
        n = s.att_n
        ok = s.att_ok
        ex = s.att_ex
        nopay = s.att_nopay
        billable = ok + ex
        pct = round(ok / n * 100) if n else 0
        out.append(
            [
                str(s),
                s.phone,
                s.group.name if s.group else '-',
                n,
                ok,
                ex,
                nopay,
                billable,
                f'{pct}%',
            ]
        )
    return out


def _students_scope_qs_for_attendance(one_student: Student | None, export_group: Group | None):
    qs = _annotated_students_export_qs()
    if one_student:
        return qs.filter(pk=one_student.pk)
    if export_group:
        return qs.filter(group=export_group)
    return qs


def _payment_rows(qs):
    rows = []
    for i, p in enumerate(qs, 1):
        rows.append([
            i,
            p.created_at.strftime('%d.%m.%Y'),
            str(p.student),
            p.group.name if p.group_id else (p.student.group.name if p.student.group else '-'),
            p.course.name if p.course_id else '-',
            p.get_payment_method_display() if getattr(p, 'payment_method', None) else "Naqd pul",
            p.get_payment_type_display(),
            int(p.amount),
            p.note or '-',
            str(p.received_by) if p.received_by else '-',
        ])
    return rows


def _payment_totals_from_qs(qs):
    """Kirim / qaytarish / chegirma va sof jami (kirim - qaytarish - chegirma)."""
    inc = int(qs.filter(payment_type=Payment.INCOME).aggregate(s=Sum('amount'))['s'] or 0)
    ref = int(qs.filter(payment_type=Payment.REFUND).aggregate(s=Sum('amount'))['s'] or 0)
    dis = int(qs.filter(payment_type=Payment.DISCOUNT).aggregate(s=Sum('amount'))['s'] or 0)
    return {'income': inc, 'refund': ref, 'discount': dis, 'net': inc - ref - dis}


def _append_xlsx_payment_totals(ws, row: int, qs, last_col: int, grid) -> int:
    """To'lovlar jadvali ostiga jami qatorlar (summa ustuni: last_col - 2)."""
    from openpyxl.styles import Alignment, Font

    totals = _payment_totals_from_qs(qs)
    lines = [
        ("Jami kirim (so'm)", totals['income'], False),
        ("Jami qaytarish (so'm)", totals['refund'], False),
        ("Jami chegirma (so'm)", totals['discount'], False),
        ("Sof jami (so'm)", totals['net'], True),
    ]
    sum_col = last_col - 2
    label_col = last_col - 3
    r = row
    for label, val, is_net in lines:
        f = Font(bold=True, size=11 if is_net else 10)
        for c in range(1, label_col):
            x = ws.cell(row=r, column=c, value='')
            x.border = grid
        c_lab = ws.cell(row=r, column=label_col, value=label)
        c_lab.font = f
        c_lab.border = grid
        c_lab.alignment = Alignment(horizontal='right', vertical='center')
        c_sum = ws.cell(row=r, column=sum_col, value=val)
        c_sum.font = f
        c_sum.number_format = '#,##0'
        c_sum.border = grid
        for c in range(sum_col + 1, last_col + 1):
            x = ws.cell(row=r, column=c, value='')
            x.border = grid
        r += 1
    return r


def _attendance_holat_tolov(a: Attendance) -> tuple[str, str]:
    if a.is_present:
        return 'Keldi', "Ha"
    if a.excused_absence:
        return 'Sababli kelmadi', "Ha"
    return 'Sababsiz kelmadi', "Yo'q"


def _one_student_attendance_bundle(student: Student):
    """
    Bitta talaba: jami yozuvlar, kelgan, sababli / sababsiz kelmagan, to'lov kunlari, foiz.
    """
    recs = list(Attendance.objects.filter(student=student).order_by('-date'))
    n = len(recs)
    ok = sum(1 for a in recs if a.is_present)
    excused = sum(1 for a in recs if (not a.is_present) and a.excused_absence)
    no_pay = sum(1 for a in recs if (not a.is_present) and (not a.excused_absence))
    billable = ok + excused
    pct = round(ok / n * 100) if n else 0
    detail = []
    for a in recs:
        holat, pay = _attendance_holat_tolov(a)
        detail.append([a.date.strftime('%d.%m.%Y'), holat, pay, a.note or '-'])
    return n, ok, excused, no_pay, billable, pct, detail


def _all_attendance_detail_rows(one_student: Student | None = None, export_group: Group | None = None):
    """Davomat yozuvlari: butun tizim, guruh yoki bitta talaba bo‘yicha."""
    rows = []
    qs = Attendance.objects.select_related('student', 'student__group').order_by(
        'student__last_name', 'student__first_name', '-date'
    )
    if one_student:
        qs = qs.filter(student=one_student)
    elif export_group:
        qs = qs.filter(student__group=export_group)
    for a in qs:
        s = a.student
        holat, pay = _attendance_holat_tolov(a)
        rows.append(
            [
                str(s),
                s.phone,
                s.group.name if s.group else '-',
                a.date.strftime('%d.%m.%Y'),
                holat,
                pay,
                a.note or '-',
            ]
        )
    return rows


def _export_attendance_csv_section(w, one_student, export_group):
    """CSV fayl oxiriga yo'qlama bo'limi."""
    w.writerow([])
    w.writerow(["Yo'qlama / davomat"])
    if one_student:
        n, ok, excused, no_pay, billable, pct, detail = _one_student_attendance_bundle(one_student)
        w.writerow(['Jami belgilangan kunlar (tizimdagi yozuvlar)', n])
        w.writerow(['Shundan kelgan', ok])
        w.writerow(["Sababli kelmagan (to'lov bor)", excused])
        w.writerow(["Sababsiz kelmagan (to'lov yo'q)", no_pay])
        w.writerow(["To'lov asosidagi kunlar (keldi + sababli)", billable])
        w.writerow(['Kelish foizi (%)', pct])
        w.writerow([])
        w.writerow(['Sana', 'Holat', "To'lov (Ha/Yo'q)", 'Izoh'])
        for drow in detail:
            w.writerow(drow)
        if not detail:
            w.writerow(["Hali davomat yozuvi yo'q."])
        return

    if export_group:
        w.writerow(['Guruh (filtr)', str(export_group)])
        w.writerow([])

    w.writerow(
        [
            'Talaba',
            'Telefon',
            'Guruh',
            'Jami',
            'Keldi',
            'Sababli kelmadi',
            'Sababsiz kelmadi',
            "To'lov kunlari",
            'Foiz',
        ]
    )
    scope_qs = _students_scope_qs_for_attendance(one_student, export_group)
    for row in _attendance_summary_rows_for_students_qs(scope_qs):
        r = list(row)
        r[1] = f'\t{r[1]}'
        w.writerow(r)
    w.writerow([])
    w.writerow(['Batafsil (sana + sababli/sababsiz + to‘lov)'])
    w.writerow(['Talaba', 'Telefon', 'Guruh', 'Sana', 'Holat', "To'lov", 'Izoh'])
    detail_all = _all_attendance_detail_rows(one_student, export_group)
    for row in detail_all:
        r = list(row)
        r[1] = f'\t{r[1]}'
        w.writerow(r)
    if not detail_all:
        w.writerow(["Batafsil uchun yozuv yo'q."])


def _write_davomat_sheet(wb, one_student, export_group):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style='thin', color='D1D5DB')
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws = wb.create_sheet(title="Yo'qlama")

    if one_student:
        last_col = 4
        r = 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        tc = ws.cell(row=r, column=1, value=f"Yo'qlama / davomat: {one_student}")
        tc.font = Font(bold=True, size=14)
        tc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 22
        r += 2

        n, ok, excused, no_pay, billable, pct, detail = _one_student_attendance_bundle(one_student)
        for label, val in [
            ('Jami belgilangan kunlar (tizimdagi yozuvlar)', n),
            ('Shundan kelgan', ok),
            ("Sababli kelmagan (to'lov bor)", excused),
            ("Sababsiz kelmagan (to'lov yo'q)", no_pay),
            ("To'lov asosidagi kunlar (keldi + sababli)", billable),
            ('Kelish foizi (%)', pct),
        ]:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            r += 1
        r += 1

        hdrs = ['Sana', 'Holat', "To'lov (Ha/Yo'q)", 'Izoh']
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = Font(bold=True)
            cell.border = grid
            cell.fill = PatternFill('solid', fgColor='E0E7FF')
        r += 1
        for drow in detail:
            for c, val in enumerate(drow, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = grid
            r += 1
        if not detail:
            ws.cell(row=r, column=1, value="Hali davomat yozuvi yo'q.")
            r += 1

        _xlsx_autofit_columns(ws, last_col)
        return

    last_col = 9
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    _dav_title = "Yo'qlama / davomat — barcha talabalar (xulosa)"
    if export_group:
        _dav_title = f"Yo'qlama / davomat — {export_group.name}"
    tc = ws.cell(row=r, column=1, value=_dav_title)
    tc.font = Font(bold=True, size=14)
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 22
    r += 2

    hint = ws.cell(
        row=r,
        column=1,
        value=(
            'Xulosa: keldi, sababli kelmadi, sababsiz kelmadi, to‘lov kunlari. '
            'Pastda sana bilan batafsil — «Holat» da sababli yoki sababsiz yoziladi.'
        ),
    )
    hint.font = Font(italic=True, size=10, color='64748B')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    r += 2

    hdrs = [
        'Talaba',
        'Telefon',
        'Guruh',
        'Jami',
        'Keldi',
        'Sababli kelmadi',
        'Sababsiz kelmadi',
        "To'lov kunlari",
        'Foiz',
    ]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True)
        cell.border = grid
        cell.fill = PatternFill('solid', fgColor='E0E7FF')
    r += 1

    _scope_qs = _students_scope_qs_for_attendance(one_student, export_group)
    for row_vals in _attendance_summary_rows_for_students_qs(_scope_qs):
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = grid
            if c == 2:
                cell.number_format = '@'
        r += 1

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    sub = ws.cell(
        row=r,
        column=1,
        value='Batafsil: har bir sana — holat (sababli / sababsiz) va to‘lov',
    )
    sub.font = Font(bold=True, size=12)
    sub.alignment = Alignment(horizontal='center', vertical='center')
    r += 1

    hint2 = ws.cell(
        row=r,
        column=1,
        value='Tartib: talaba (A–Z), ichida sana yangisidan eskisiga.',
    )
    hint2.font = Font(italic=True, size=10, color='64748B')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    r += 2

    hdrs_detail = ['Talaba', 'Telefon', 'Guruh', 'Sana', 'Holat', "To'lov", 'Izoh']
    for c, h in enumerate(hdrs_detail, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True)
        cell.border = grid
        cell.fill = PatternFill('solid', fgColor='DCFCE7')
    r += 1

    detail_all = _all_attendance_detail_rows(one_student, export_group)
    for row_vals in detail_all:
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = grid
            if c == 2:
                cell.number_format = '@'
        r += 1
    if not detail_all:
        ws.cell(row=r, column=1, value="Hali batafsil yozuv yo'q (davomat kiritilmagan).")
        r += 1

    _xlsx_autofit_columns(ws, last_col)


def _parse_export_day(s: str | None) -> date:
    if not s:
        return date.today()
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return date.today()


def _parse_export_week_range(s: str | None) -> tuple[date, date]:
    if not s or '-W' not in str(s):
        t = date.today()
        y, w, _ = t.isocalendar()
        d0 = date.fromisocalendar(y, w, 1)
        d1 = date.fromisocalendar(y, w, 7)
        return d0, d1
    try:
        y, rest = str(s).split('-W', 1)
        wn = int(rest)
        yn = int(y)
        d0 = date.fromisocalendar(yn, wn, 1)
        d1 = date.fromisocalendar(yn, wn, 7)
        return d0, d1
    except ValueError:
        t = date.today()
        y, w, _ = t.isocalendar()
        return date.fromisocalendar(y, w, 1), date.fromisocalendar(y, w, 7)


def _parse_export_month_range(s: str | None) -> tuple[date, date]:
    if not s:
        t = date.today()
        s = f'{t.year}-{t.month:02d}'
    try:
        y, m = map(int, s.split('-')[:2])
        last = calendar.monthrange(y, m)[1]
        return date(y, m, 1), date(y, m, last)
    except (ValueError, IndexError):
        t = date.today()
        last = calendar.monthrange(t.year, t.month)[1]
        return date(t.year, t.month, 1), date(t.year, t.month, last)


def _resolve_period_students_qs(scope: str, group_raw: str, student_raw: str):
    """Faqat admin eksport — barcha faol talabalar bazasi."""
    base = Student.objects.filter(is_active=True).select_related('group')
    g_obj = None
    s_obj = None
    if scope == 'guruh' and group_raw.isdigit():
        g_obj = get_object_or_404(Group, pk=int(group_raw), is_active=True)
        return base.filter(group=g_obj), g_obj, None
    if scope == 'talaba' and student_raw.isdigit():
        s_obj = get_object_or_404(Student, pk=int(student_raw), is_active=True)
        return base.filter(pk=s_obj.pk), None, s_obj
    return base, None, None


def _scope_label_period(scope: str, group_obj, student_obj) -> str:
    if scope == 'guruh' and group_obj:
        return f"Guruh: {group_obj.name}"
    if scope == 'talaba' and student_obj:
        return f"Talaba: {student_obj.full_name}"
    return "Butun o'quv markaz"


def _metrics_period_block(students_qs, start: date, end: date, payment_group: Group | None = None) -> dict:
    pks = list(students_qs.values_list('pk', flat=True))
    if not pks:
        return {
            'student_count': 0,
            'income': 0,
            'payments_n': 0,
            'att_n': 0,
            'att_present': 0,
            'att_billable': 0,
            'att_pct': 0,
            'debtors': 0,
        }
    pay_base = Payment.objects.filter(
        student_id__in=pks,
        created_at__date__gte=start,
        created_at__date__lte=end,
    )
    if payment_group is not None:
        # Yangi yozuvlar: payment.group; eski yozuvlar: student.group fallback
        pay_base = pay_base.filter(
            Q(group=payment_group) | Q(group__isnull=True, student__group=payment_group)
        )
    income = int(pay_base.filter(payment_type=Payment.INCOME).aggregate(s=Sum('amount'))['s'] or 0)
    payments_n = pay_base.count()
    att = Attendance.objects.filter(student_id__in=pks, date__gte=start, date__lte=end)
    att_n = att.count()
    att_present = att.filter(is_present=True).count()
    att_billable = att.filter(
        Q(is_present=True) | Q(is_present=False, excused_absence=True)
    ).count()
    att_pct = round(att_present / att_n * 100) if att_n else 0
    debtors = students_qs.filter(balance__lt=0).count()
    return {
        'student_count': len(set(pks)),
        'income': income,
        'payments_n': payments_n,
        'att_n': att_n,
        'att_present': att_present,
        'att_billable': att_billable,
        'att_pct': att_pct,
        'debtors': debtors,
    }


def _teacher_group_revenue_rows(students_qs, pay_qs):
    """
    Filtr bo'yicha talabalar: har bir guruh uchun o'qituvchi, talabalar soni,
    davrdagi kirim / qaytarish / chegirma / sof jami (to'lovlar pay_qs dan).
    """
    from collections import Counter, defaultdict

    scope_pks = set(students_qs.values_list('pk', flat=True))
    if not scope_pks:
        return []

    group_students = Counter(
        Student.objects.filter(pk__in=scope_pks, group_id__isnull=False).values_list('group_id', flat=True)
    )

    sums = defaultdict(lambda: {'income': 0, 'refund': 0, 'discount': 0})
    meta = {}
    for p in pay_qs.select_related('group__teacher', 'student__group', 'student__group__teacher'):
        g = p.group if p.group_id else p.student.group
        gid = g.pk if g else 0
        if gid not in meta:
            if g:
                meta[gid] = {
                    'teacher': str(g.teacher) if g.teacher_id else '—',
                    'teacher_id': g.teacher_id,
                    'group': g.name,
                }
            else:
                meta[gid] = {'teacher': '—', 'teacher_id': None, 'group': 'Guruhsiz talabalar'}
        amt = int(p.amount)
        if p.payment_type == Payment.INCOME:
            sums[gid]['income'] += amt
        elif p.payment_type == Payment.REFUND:
            sums[gid]['refund'] += amt
        elif p.payment_type == Payment.DISCOUNT:
            sums[gid]['discount'] += amt

    all_gids = set(group_students.keys()) | set(sums.keys()) | set(meta.keys())
    for gid in list(all_gids):
        if gid and gid not in meta:
            g = Group.objects.filter(pk=gid).select_related('teacher').first()
            if g:
                meta[gid] = {
                    'teacher': str(g.teacher) if g.teacher_id else '—',
                    'teacher_id': g.teacher_id,
                    'group': g.name,
                }

    rows = []
    for gid in sorted(
        all_gids,
        key=lambda x: (
            (meta.get(x) or {}).get('teacher') or '',
            (meta.get(x) or {}).get('group') or '',
        ),
    ):
        m = meta.get(gid) or {'teacher': '—', 'teacher_id': None, 'group': '—'}
        s = sums.get(gid, {'income': 0, 'refund': 0, 'discount': 0})
        net = s['income'] - s['refund'] - s['discount']
        rows.append(
            {
                'teacher': m['teacher'],
                'teacher_id': m['teacher_id'],
                'group': m['group'],
                'students_n': group_students.get(gid, 0) if gid else Student.objects.filter(
                    pk__in=scope_pks, group_id__isnull=True
                ).count(),
                'income': s['income'],
                'refund': s['refund'],
                'discount': s['discount'],
                'net': net,
            }
        )
    return rows


def _write_teacher_group_revenue_sheet(wb, students_qs, pay_qs, period_label: str):
    """Excel varag'i: o'qituvchi / guruh / pul (har guruh alohida, o'qituvchi bo'yicha jami qatorlari)."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style='thin', color='D1D5DB')
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='EEF2FF')
    sub_fill = PatternFill('solid', fgColor='F3F4F6')

    ws = wb.create_sheet(title='Oqituvchi va guruh')
    last_col = 7
    rr = 1
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=last_col)
    tc = ws.cell(row=rr, column=1, value=f"O'qituvchi va guruhlar bo'yicha pul — {period_label}")
    tc.font = Font(bold=True, size=12)
    tc.alignment = Alignment(horizontal='center')
    rr += 2

    headers = [
        "O'qituvchi",
        'Guruh',
        'Talabalar (filtr)',
        "Kirim (so'm)",
        "Qaytarish (so'm)",
        "Chegirma (so'm)",
        "Sof jami (so'm)",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=rr, column=c, value=h)
        cell.font = Font(bold=True)
        cell.border = grid
        cell.fill = hdr_fill
    rr += 1

    rows = _teacher_group_revenue_rows(students_qs, pay_qs)
    if not rows:
        ws.cell(row=rr, column=1, value="Ma'lumot yo'q.")
        rr += 1
        _xlsx_autofit_columns(ws, last_col)
        return

    from itertools import groupby

    def _tid(r):
        return r['teacher_id'] if r['teacher_id'] is not None else -1

    rows.sort(key=lambda r: (_tid(r), r['group']))
    for tid, block in groupby(rows, key=_tid):
        block = list(block)
        for r in block:
            vals = [
                r['teacher'],
                r['group'],
                r['students_n'],
                r['income'],
                r['refund'],
                r['discount'],
                r['net'],
            ]
            for c, val in enumerate(vals, 1):
                cell = ws.cell(row=rr, column=c, value=val)
                cell.border = grid
                if c >= 4:
                    cell.number_format = '#,##0'
            rr += 1

        sub_inc = sum(x['income'] for x in block)
        sub_ref = sum(x['refund'] for x in block)
        sub_dis = sum(x['discount'] for x in block)
        sub_net = sub_inc - sub_ref - sub_dis
        sub_teacher = block[0]['teacher'] if block else '—'
        sub_vals = [sub_teacher, ">>> Jami (shu o'qituvchi)", '', sub_inc, sub_ref, sub_dis, sub_net]
        for c, val in enumerate(sub_vals, 1):
            cell = ws.cell(row=rr, column=c, value=val)
            cell.font = Font(bold=True)
            cell.border = grid
            cell.fill = sub_fill
            if c >= 4 and val != '':
                cell.number_format = '#,##0'
        rr += 1

    gi = sum(r['income'] for r in rows)
    gr = sum(r['refund'] for r in rows)
    gd = sum(r['discount'] for r in rows)
    gn = gi - gr - gd
    for c, val in enumerate(
        ['', '', 'UMUMIY JAMI (barcha guruhlar)', gi, gr, gd, gn],
        start=1,
    ):
        cell = ws.cell(row=rr, column=c, value=val)
        cell.font = Font(bold=True, size=11)
        cell.border = grid
        cell.fill = hdr_fill
        if c >= 4:
            cell.number_format = '#,##0'

    _xlsx_autofit_columns(ws, last_col)


def _attendance_rows_period(students_qs, start: date, end: date):
    pks = list(students_qs.values_list('pk', flat=True))
    if not pks:
        return []
    qs = (
        Attendance.objects.filter(student_id__in=pks, date__gte=start, date__lte=end)
        .select_related('student', 'student__group')
        .order_by('student__last_name', 'student__first_name', '-date')
    )
    rows = []
    for a in qs:
        s = a.student
        holat, pay = _attendance_holat_tolov(a)
        rows.append(
            [
                str(s),
                s.phone,
                s.group.name if s.group else '-',
                a.date.strftime('%d.%m.%Y'),
                holat,
                pay,
                a.note or '-',
            ]
        )
    return rows


def _build_period_xlsx_response(
    start: date,
    end: date,
    kind_title: str,
    scope_label: str,
    period_range_label: str,
    metrics: dict,
    pay_qs,
    students_qs,
    filename_suffix: str,
):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style='thin', color='D1D5DB')
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws0 = wb.active
    ws0.title = 'Xulosa'

    r = 1
    ws0.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    t = ws0.cell(row=r, column=1, value=f'EduFlow — {kind_title}')
    t.font = Font(bold=True, size=14)
    t.alignment = Alignment(horizontal='center')
    r += 2

    pairs = [
        ('Davr', period_range_label),
        ('Tanlangan doira', scope_label),
        ('Talabalar (filtr bo‘yicha)', metrics['student_count']),
        ("Kirim (so'm, faqat kirim)", metrics['income']),
        ('Barcha to‘lov yozuvlari (davr)', metrics['payments_n']),
        ('Davomat yozuvlari (davr)', metrics['att_n']),
        ('Shundan «keldi»', metrics['att_present']),
        ("To'lov kuni (keldi + sababli)", metrics['att_billable']),
        ('Keldi % (yozuvlar bo‘yicha)', f"{metrics['att_pct']}%"),
        ('Qarzdorlar (joriy balans < 0)', metrics['debtors']),
    ]
    for label, val in pairs:
        ws0.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws0.cell(row=r, column=2, value=val).border = grid
        r += 1
    _xlsx_autofit_columns(ws0, 2)

    ws1 = wb.create_sheet(title="To'lovlar")
    last_col = 9
    rr = 1
    ws1.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=last_col)
    tc = ws1.cell(
        row=rr,
        column=1,
        value=f"To'lovlar — {period_range_label} ({start.isoformat()} … {end.isoformat()})",
    )
    tc.font = Font(bold=True, size=12)
    rr += 2
    headers = ['#', 'Sana', 'Talaba', 'Guruh', 'Fan', "To'lov usuli", 'Tur', "Summa (so'm)", 'Izoh', 'Kim qabul qildi']
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=rr, column=c, value=h)
        cell.font = Font(bold=True)
        cell.border = grid
        cell.fill = PatternFill('solid', fgColor='EEF2FF')
    rr += 1
    had_pay = False
    for row in _payment_rows(pay_qs):
        for c, val in enumerate(row, 1):
            cell = ws1.cell(row=rr, column=c, value=val)
            cell.border = grid
            if c == 8:
                cell.number_format = '#,##0'
        had_pay = True
        rr += 1
    if not had_pay:
        ws1.cell(row=rr, column=1, value="Tanlangan davrda to‘lov yozuvi yo‘q.")
        rr += 1
    else:
        rr += 1
    rr = _append_xlsx_payment_totals(ws1, rr, pay_qs, last_col, grid)
    _xlsx_autofit_columns(ws1, last_col)

    period_label = f"{period_range_label} ({start.isoformat()} ... {end.isoformat()})"
    _write_teacher_group_revenue_sheet(wb, students_qs, pay_qs, period_label)

    ws2 = wb.create_sheet(title="Davomat")
    rr = 1
    ws2.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=7)
    tc2 = ws2.cell(row=rr, column=1, value=f"Davomat batafsil — {period_range_label}")
    tc2.font = Font(bold=True, size=12)
    rr += 2
    hdrs2 = ['Talaba', 'Telefon', 'Guruh', 'Sana', 'Holat', "To'lov", 'Izoh']
    for c, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=rr, column=c, value=h)
        cell.font = Font(bold=True)
        cell.border = grid
        cell.fill = PatternFill('solid', fgColor='DCFCE7')
    rr += 1
    detail = _attendance_rows_period(students_qs, start, end)
    for row_vals in detail:
        for c, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=rr, column=c, value=val)
            cell.border = grid
            if c == 2:
                cell.number_format = '@'
        rr += 1
    if not detail:
        ws2.cell(row=rr, column=1, value="Tanlangan davrda davomat yozuvi yo‘q.")
        rr += 1
    _xlsx_autofit_columns(ws2, 7)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    response = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="eduflow_{filename_suffix}.xlsx"'
    return response


@login_required
def export_period_report(request):
    """Kunlik / haftalik / oylik — Excel (xulosa, to'lovlar, o'qituvchi/guruh puli, davomat)."""
    admin_redirect = _ensure_admin(request)
    if admin_redirect:
        return admin_redirect

    kind = (request.GET.get('kind') or '').strip().lower()
    if kind not in ('kunlik', 'haftalik', 'oylik'):
        return HttpResponseBadRequest('kind=kunlik, haftalik yoki oylik bo‘lishi kerak.')

    scope = (request.GET.get('scope') or 'markaz').strip()
    if scope not in ('markaz', 'guruh', 'talaba'):
        scope = 'markaz'

    group_raw = (request.GET.get('group') or '').strip()
    student_raw = (request.GET.get('student') or '').strip()

    if kind == 'kunlik':
        start = end = _parse_export_day(request.GET.get('d_date'))
        period_range_label = start.strftime('%d.%m.%Y')
        fn_suffix = f'kunlik_{start.isoformat()}'
        kind_title = 'Kunlik hisobot'
    elif kind == 'haftalik':
        start, end = _parse_export_week_range((request.GET.get('w_week') or '').strip())
        period_range_label = f"{start.strftime('%d.%m.%Y')} — {end.strftime('%d.%m.%Y')}"
        fn_suffix = f'haftalik_{start.isoformat()}_{end.isoformat()}'
        kind_title = 'Haftalik hisobot'
    else:
        start, end = _parse_export_month_range((request.GET.get('m_month') or '').strip())
        period_range_label = f"{start.strftime('%m.%Y')}"
        fn_suffix = f'oylik_{start.year}-{start.month:02d}'
        kind_title = 'Oylik hisobot'

    students_qs, grp_obj, st_obj = _resolve_period_students_qs(scope, group_raw, student_raw)
    scope_label = _scope_label_period(scope, grp_obj, st_obj)
    metrics = _metrics_period_block(students_qs, start, end, grp_obj if scope == 'guruh' else None)
    pks = list(students_qs.values_list('pk', flat=True))
    pay_qs = Payment.objects.filter(
        student_id__in=pks,
        created_at__date__gte=start,
        created_at__date__lte=end,
    ).select_related('student__group', 'received_by', 'course', 'group').order_by('-created_at')
    if scope == 'guruh' and grp_obj is not None:
        pay_qs = pay_qs.filter(Q(group=grp_obj) | Q(group__isnull=True, student__group=grp_obj))

    return _build_period_xlsx_response(
        start,
        end,
        kind_title,
        scope_label,
        period_range_label,
        metrics,
        pay_qs,
        students_qs,
        fn_suffix,
    )


@login_required
def export_report(request):
    admin_redirect = _ensure_admin(request)
    if admin_redirect:
        return admin_redirect

    fmt = (request.GET.get('format') or 'xlsx').strip().lower()
    student_raw = (request.GET.get('student') or '').strip()
    group_raw = (request.GET.get('group') or '').strip()
    qs, one_student, export_group = _payment_export_queryset(
        student_raw or None,
        group_raw or None,
    )

    if fmt == 'csv':
        return _export_payments_csv(qs, one_student, export_group)
    if fmt != 'xlsx':
        return HttpResponseBadRequest('format=xlsx yoki csv bo‘lishi kerak.')
    return _export_payments_xlsx(qs, one_student, export_group)


def _export_payments_csv(qs, one_student, export_group):
    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    suffix = _export_filename_suffix(one_student, export_group)
    response['Content-Disposition'] = f'attachment; filename="eduflow_moliya{suffix}.csv"'
    # BOM: Excel Windows UTF-8 ni to‘g‘ri o‘qishi uchun
    response.write('\ufeff')
    response.write('sep=;\r\n')
    w = csv.writer(response, delimiter=';', lineterminator='\r\n')
    w.writerow(['EduFlow - moliya hisoboti'])
    if export_group and not one_student:
        w.writerow([])
        w.writerow(['Guruh (filtr)', str(export_group)])
    if one_student:
        w.writerow([])
        w.writerow(['Talaba (filtr)', str(one_student)])
        # Boshidagi tab: Excel raqamni ilmiy ko‘rinishga aylantirmasligi uchun
        w.writerow(['Telefon', f'\t{one_student.phone}'])
        w.writerow(['Guruh', one_student.group.name if one_student.group else '-'])
        w.writerow(["Balans (so'm)", int(one_student.balance)])
        w.writerow([])
    w.writerow(['#', 'Sana', 'Talaba', 'Guruh', 'Fan', "To'lov usuli", 'Tur', "Summa (so'm)", 'Izoh', 'Kim qabul qildi'])
    for row in _payment_rows(qs):
        w.writerow(row)
    t = _payment_totals_from_qs(qs)
    w.writerow([])
    w.writerow(['', '', '', '', '', "Jami kirim (so'm)", t['income'], '', ''])
    w.writerow(['', '', '', '', '', "Jami qaytarish (so'm)", t['refund'], '', ''])
    w.writerow(['', '', '', '', '', "Jami chegirma (so'm)", t['discount'], '', ''])
    w.writerow(['', '', '', '', '', "Sof jami (so'm)", t['net'], '', ''])
    _export_attendance_csv_section(w, one_student, export_group)
    return response


def _xlsx_autofit_columns(ws, max_col: int, min_w: float = 12.0, cap_w: float = 52.0):
    """Ustun kengligini mazmun bo‘yicha sozlaydi (Excelda kesilmasin)."""
    from openpyxl.utils import get_column_letter

    for col in range(1, max_col + 1):
        best = min_w
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            ln = len(str(val))
            if ln + 2 > best:
                best = min(cap_w, ln + 2)
        ws.column_dimensions[get_column_letter(col)].width = best


def _students_qs_for_moliya_export(qs, one_student, export_group):
    """Moliya Excel: o'qituvchi/guruh varag'i uchun talaba filtri (tranzaksiya ro'yxati bilan mos)."""
    if one_student:
        return Student.objects.filter(pk=one_student.pk)
    if export_group:
        return Student.objects.filter(group=export_group, is_active=True)
    return Student.objects.filter(is_active=True)


def _export_payments_xlsx(qs, one_student, export_group):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    last_col = 10
    thin = Side(style='thin', color='D1D5DB')
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Moliya'

    r = 1
    # Uzun tire o‘rniga oddiy tire — Excelda kodlash chalkashligi bo‘lmasin
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    title_cell = ws.cell(row=r, column=1, value='EduFlow - moliya hisoboti')
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 22
    r += 1

    if export_group and not one_student:
        r += 1
        ws.cell(row=r, column=1, value='Guruh (filtr)')
        ws.cell(row=r, column=2, value=str(export_group))
        r += 1

    if one_student:
        r += 1
        ws.cell(row=r, column=1, value='Talaba (filtr)')
        ws.cell(row=r, column=2, value=str(one_student))
        r += 1
        ws.cell(row=r, column=1, value='Telefon')
        ph = ws.cell(row=r, column=2, value=str(one_student.phone))
        ph.number_format = '@'
        r += 1
        ws.cell(row=r, column=1, value='Guruh')
        ws.cell(row=r, column=2, value=one_student.group.name if one_student.group else '-')
        r += 1
        ws.cell(row=r, column=1, value="Balans (so'm)")
        bal = ws.cell(row=r, column=2, value=int(one_student.balance))
        bal.number_format = '#,##0'
        r += 1
        r += 1

    headers = ['#', 'Sana', 'Talaba', 'Guruh', 'Fan', "To'lov usuli", 'Tur', "Summa (so'm)", 'Izoh', 'Kim qabul qildi']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True)
        cell.border = grid
        cell.fill = PatternFill('solid', fgColor='EEF2FF')
    r += 1

    data_row0 = r
    for row in _payment_rows(qs):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = grid
            if c == 1:
                cell.alignment = Alignment(horizontal='center')
            if c == 8:
                cell.number_format = '#,##0'
        r += 1

    if r > data_row0:
        r += 1
    r = _append_xlsx_payment_totals(ws, r, qs, last_col, grid)

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    note = ws.cell(
        row=r,
        column=1,
        value="Yo'qlama: alohida varaqda sababli / sababsiz kelmagan kunlar va to'lov (Ha/Yo'q).",
    )
    note.font = Font(italic=True, size=10, color='64748B')
    note.alignment = Alignment(wrap_text=True, vertical='center')

    _xlsx_autofit_columns(ws, last_col)

    st_qs = _students_qs_for_moliya_export(qs, one_student, export_group)
    _write_teacher_group_revenue_sheet(wb, st_qs, qs, 'Moliya eksporti')

    _write_davomat_sheet(wb, one_student, export_group)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    suffix = _export_filename_suffix(one_student, export_group)
    response['Content-Disposition'] = f'attachment; filename="eduflow_moliya{suffix}.xlsx"'
    return response
